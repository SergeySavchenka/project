from PyQt6.QtWidgets import *
from PyQt6.QtCore import Qt
import mysql.connector
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class AdminDatabase(QWidget):
    def __init__(self):
        self.connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="inform_sys"
        )

        self.cursor = self.connection.cursor()
        super().__init__()

        self.setWindowTitle("Данные")
        self.setGeometry(0, 0, 600, 400)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.label_info = QLabel("Выберите таблицу для отображения данных:")
        self.table_selector = QComboBox()
        self.populate_table_combo()

        choice_layout = QHBoxLayout()
        choice_layout.addWidget(self.label_info)
        choice_layout.addWidget(self.table_selector)
        layout.addLayout(choice_layout)

        self.add_button = QPushButton("Добавить запись")
        self.add_button.clicked.connect(self.show_add_data_dialog)
        self.delete_button = QPushButton("Удалить запись")
        self.delete_button.clicked.connect(self.show_delete_data_dialog)
        self.edit_button = QPushButton('Редактировать запись')
        self.edit_button.clicked.connect(self.show_edit_data_dialog)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.add_button)
        button_layout.addWidget(self.delete_button)
        button_layout.addWidget(self.edit_button)
        layout.addLayout(button_layout)

        self.data_table = QTableWidget()
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.data_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.data_table.verticalHeader().setVisible(False)

        self.btn_show_data = QPushButton("Показать данные", self)
        self.btn_show_data.clicked.connect(self.show_data)

        layout.addWidget(self.label_info)
        layout.addWidget(self.table_selector)
        layout.addWidget(self.data_table)
        layout.addWidget(self.btn_show_data)

        self.setLayout(layout)

        close_button = QPushButton('Закрыть окно')
        close_button.clicked.connect(self.close)
        layout.addWidget(close_button)
        self.center()

    def center(self):
        screen_geometry = QApplication.primaryScreen().geometry()
        window_geometry = self.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.move(window_geometry.topLeft())

    def show_data(self):
        # Очистка таблицы перед отображением новых данных
        self.data_table.clear()
        self.data_table.setRowCount(0)

        # Выбор таблицы из выпадающего списка
        selected_table = self.table_selector.currentText()

        self.cursor.execute(f"SELECT * FROM {selected_table}")
        data = self.cursor.fetchall()

        # Если есть данные
        if data:
            # Определение числа столбцов и строк
            num_rows = len(data)
            num_cols = len(data[0])

            # Установка размера таблицы
            self.data_table.setRowCount(num_rows)
            self.data_table.setColumnCount(num_cols)

            # Установка заголовков столбцов
            column_headers = [description[0] for description in self.cursor.description]
            self.data_table.setHorizontalHeaderLabels(column_headers)

            # Заполнение ячеек таблицы данными
            for row in range(num_rows):
                for col in range(num_cols):
                    item = QTableWidgetItem(str(data[row][col]))
                    self.data_table.setItem(row, col, item)

    def populate_table_combo(self):
        # Получение списка таблиц из базы данных
        self.cursor.execute("SHOW TABLES")
        tables = self.cursor.fetchall()

        # Заполнение выпадающего списка
        for table in tables:
            self.table_selector.addItem(table[0])

    def show_add_data_dialog(self):
        # Определение выбранной таблицы
        selected_table = self.table_selector.currentText()

        # Получение названий столбцов для выбранной таблицы
        self.cursor.execute(f"DESCRIBE {selected_table}")
        columns = [column[0] for column in self.cursor.fetchall()]

        # Создание диалогового окна с полями ввода для каждого столбца и кнопкой "Добавить"
        form_layout = QFormLayout()
        input_fields = []

        for column in columns:
            label = QLabel(column)
            input_field = QLineEdit()
            form_layout.addRow(label, input_field)
            input_fields.append(input_field)

        dialog = QDialog(self)
        dialog.setWindowTitle('Ввод данных')
        dialog.setLayout(form_layout)

        # Обработчик нажатия кнопки "Добавить"
        def add_data():
            try:
                # Получение значений из полей ввода
                values = [field.text() for field in input_fields]

                # Убедитесь, что количество введенных значений соответствует количеству столбцов
                if len(values) == len(columns):
                    # Формирование SQL-запроса для добавления данных
                    query = f"INSERT INTO {selected_table} ({', '.join(columns)}) VALUES ({', '.join(['%s'] * len(columns))})"
                    self.cursor.execute(query, values)
                    self.connection.commit()

                    QMessageBox.information(self, "Успешно", "Данные добавлены успешно!")
                    self.show_data()  # Обновление отображения данных
                    dialog.accept()  # Закрытие диалогового окна после успешного добавления данных

                else:
                    QMessageBox.warning(self, 'Предупреждение об ошибке', 'Количество введенных значений не совпадает с количеством столбцов')
            except:
                QMessageBox.warning(self, 'Предупреждение об ошибке', 'Ошибка!')

        add_button = QPushButton('Добавить')
        form_layout.addRow(add_button)
        add_button.clicked.connect(add_data)

        # Отображение диалогового окна
        result = dialog.exec()

    def show_delete_data_dialog(self):
        # Определение выбранной таблицы
        selected_table = self.table_selector.currentText()

        # Создание диалогового окна с полем ввода для значения id и кнопкой "Удалить"
        form_layout = QFormLayout()
        id_input = QLineEdit()
        delete_button = QPushButton('Удалить')
        form_layout.addRow("Введите ID для удаления данных:", id_input)
        form_layout.addRow(delete_button)

        dialog = QDialog(self)
        dialog.setWindowTitle('Удаление данных')
        dialog.setLayout(form_layout)

        # Обработчик нажатия кнопки "Удалить"
        def delete_data():
            self.cursor.execute(f"DESCRIBE {selected_table}")
            id_column = [column[0] for column in self.cursor.fetchall()][0]

            # Получение значения из поля ввода
            id_value = id_input.text()

            # Формирование SQL-запроса для удаления данных
            query = f"DELETE FROM {selected_table} WHERE {id_column} = {id_value}"
            self.cursor.execute(query)
            self.connection.commit()

            QMessageBox.information(self, "Успешно", "Данные удалены успешно!")
            self.show_data()  # Обновление отображения данных
            dialog.accept()  # Закрытие диалогового окна после успешного удаления данных

        delete_button.clicked.connect(delete_data)

        # Отображение диалогового окна
        result = dialog.exec()

    def show_edit_data_dialog(self):
        selected_data = self.get_selected_data()

        if selected_data:
            edit_dialog = QDialog(self)
            edit_dialog.setWindowTitle('Редактировать запись')

            layout = QFormLayout()

            # Примеры полей редактирования (замените их на необходимые поля)
            input_fields = []
            for key, value in selected_data.items():
                label = QLabel(key.capitalize())
                edit_line = QLineEdit(self)
                if layout.rowCount() == 0:
                    edit_line.setEnabled(False)
                edit_line.setText(str(value))
                layout.addRow(label, edit_line)
                input_fields.append(edit_line)

            save_button = QPushButton('Сохранить')
            save_button.clicked.connect(lambda: self.save_edited_data(selected_data, input_fields))
            layout.addRow(save_button)

            edit_dialog.setLayout(layout)

            result = edit_dialog.exec()

        else:
            QMessageBox.warning(self, 'Предупреждение об ошибке', 'Выберите запись для редактирования')

    def get_selected_data(self):
        selected_items = self.data_table.selectedItems()
        if selected_items:
            selected_row = selected_items[0].row()
            selected_data = {}
            for col in range(self.data_table.columnCount()):
                header = self.data_table.horizontalHeaderItem(col).text()
                item = self.data_table.item(selected_row, col)
                selected_data[header] = item.text()
            return selected_data
        return None

    def save_edited_data(self, selected_data, input_fields):
        # Реализуйте сохранение отредактированных данных в базе данных
        try:
            selected_id = selected_data.get(
                "id")  # Предположим, что у вас есть столбец с уникальными идентификаторами "id"
            if not selected_id:
                raise ValueError("Отсутствует уникальный идентификатор для редактирования.")

            update_values = {key: field.text() for key, field in zip(selected_data.keys(), input_fields)}

            set_clause = ", ".join([f"{key} = %s" for key in update_values.keys()])
            query = f"UPDATE {self.table_selector.currentText()} SET {set_clause} WHERE id = {selected_id}"

            self.cursor.execute(query, tuple(update_values.values()))
            self.connection.commit()

            print("Данные успешно отредактированы")
            self.show_data()  # Обновление отображения данных
        except:
            QMessageBox.warning(self, 'Предупреждение об ошибке', 'Ошибка при редактировании данных')


class AdminExcel(QWidget):
    def __init__(self):
        self.connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="inform_sys"
        )

        self.cursor = self.connection.cursor()
        super().__init__()

        self.setWindowTitle("Excel")
        self.setGeometry(0, 0, 300, 200)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.label_info = QLabel("Выберите варианты действия:")
        self.label_info.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        self.label_info.adjustSize()

        drivers_button = QPushButton("Собрать информацию о маршрутах водителя")
        drivers_button.clicked.connect(self.drivers_to_excel)
        autos_button = QPushButton('Собрать информацию о маршрутах машины')
        autos_button.clicked.connect(self.autos_to_excel)
        med_button = QPushButton('Собрать информацию о прохождении мед. осмотра')
        med_button.clicked.connect(self.medicine_to_excel)

        buttons_layout = QVBoxLayout()
        buttons_layout.addWidget(drivers_button)
        buttons_layout.addWidget(autos_button)
        buttons_layout.addWidget(med_button)

        layout.addWidget(self.label_info)
        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        close_button = QPushButton('Закрыть окно')
        close_button.clicked.connect(self.close)
        layout.addWidget(close_button)
        self.center()

    def center(self):
        screen_geometry = QApplication.primaryScreen().geometry()
        window_geometry = self.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.move(window_geometry.topLeft())

    def drivers_to_excel(self):
        def export():
            selected_table = drivers_cb.currentText()
            self.cursor.execute(
                f'select rout_id, concat(mark, " ", model), concat(female, " ", name), departure_date, arrival_date, distance, destination from routs '
                f'inner join driver on driver.driver_id = routs.driver_id '
                f'inner join auto on auto.auto_id = routs.auto_id '
                f'where driver.driver_id = {selected_table.split()[0]}')
            data = self.cursor.fetchall()
            headers = ["rout id", "auto", "driver", "departure_date", "arrival_date", "distance", "destination"]
            file_name = f'driver_{selected_table.split()[2]}.xlsx'
            self.export_to_excel(data, headers, file_name)
            dialog.close()

        dialog = QDialog(self)
        dialog.setWindowTitle('Маршруты')

        drivers_cb = QComboBox()
        self.cursor.execute('select driver_id, name, female from driver')
        driver_info = [list(sub_list) for sub_list in self.cursor.fetchall()]
        drivers_cb.addItems([f'{str(item[0])} - {item[2]} {item[1]}' for item in driver_info])

        show_table = QPushButton('Вывести в excel')
        drivers_label = QLabel('Выберите водителя:')
        dr_layout = QHBoxLayout()
        dr_layout.addWidget(drivers_label)
        dr_layout.addWidget(drivers_cb)
        show_table.clicked.connect(export)
        layout = QVBoxLayout()
        layout.addLayout(dr_layout)
        layout.addWidget(show_table)
        dialog.setLayout(layout)
        result = dialog.exec()

    def autos_to_excel(self):
        def export():
            selected_table = auto_cb.currentText()
            self.cursor.execute(
                f'select rout_id, concat(mark, " ", model), concat(female, " ", name), departure_date, arrival_date, distance, destination from routs '
                f'inner join auto on auto.auto_id = routs.auto_id '
                f'inner join driver on driver.driver_id = routs.driver_id '
                f'where auto.auto_id = {selected_table.split()[0]}')
            data = self.cursor.fetchall()
            headers = ["rout id", "auto", "driver", "departure_date", "arrival_date", "distance", "destination"]
            file_name = f'auto_{selected_table.split()[0]}.xlsx'
            self.export_to_excel(data, headers, file_name)
            dialog.close()

        dialog = QDialog(self)
        dialog.setWindowTitle('Маршруты')

        auto_cb = QComboBox()
        self.cursor.execute('select auto_id, mark, model from auto')
        car_info = [list(sub_list) for sub_list in self.cursor.fetchall()]
        auto_cb.addItems([f'{str(item[0])} - {item[1]} {item[2]}' for item in car_info])

        show_table = QPushButton('Вывести в excel')
        autos_label = QLabel('Выберите автомобиль:')
        auto_layout = QHBoxLayout()
        auto_layout.addWidget(autos_label)
        auto_layout.addWidget(auto_cb)
        show_table.clicked.connect(export)
        layout = QVBoxLayout()
        layout.addLayout(auto_layout)
        layout.addWidget(show_table)
        dialog.setLayout(layout)
        result = dialog.exec()

    def medicine_to_excel(self):
        self.cursor.execute(
            f'select driver_id, name, female, drivers_license_number, last_med_exam_date, '
            f'date_add(last_med_exam_date , INTERVAL 90 DAY) as next_med_exam_date, '
            f'if(datediff(now(), date_add(last_med_exam_date , INTERVAL 90 DAY)) > 0, "true", "false") as completed_status '
            f'from driver '
                            )
        data = self.cursor.fetchall()
        headers = ["driver_id", "name", "female", "drivers_license_number", "last_med_exam_date", "next_med_exam_date", "completed_status"]
        file_name = 'med_exams.xlsx'
        self.export_to_excel(data, headers, file_name)

    def export_to_excel(self, data, headers, file_name):
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            for col_idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(col_idx)
                cell = f"{column_letter}1"
                worksheet[cell] = header
                worksheet[cell].alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    column_letter = get_column_letter(col_idx)
                    cell = f"{column_letter}{row_idx}"
                    worksheet[cell] = value

            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            workbook.save(file_name)
            QMessageBox.information(self, 'Успешно', 'Отчёт создан!')
        except:
            QMessageBox.warning(self, 'Ошибка', 'Ошибка при экспорте данных')
