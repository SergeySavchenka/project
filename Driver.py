from PyQt6.QtWidgets import *
import mysql.connector
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class DriverWindow(QWidget):
    def __init__(self, driver_id):
        super().__init__()
        self.driver_id = int(driver_id)

        self.connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="root",
            database="inform_sys"
        )

        self.cursor = self.connection.cursor()

        self.setWindowTitle("Карточка водителя")
        self.setGeometry(0, 0, 200, 300)

        self.init_ui()

    def init_ui(self):
        self.cursor.execute(f"DESCRIBE driver")
        columns = [column[0] for column in self.cursor.fetchall()][:-1]

        self.cursor.execute(f"SELECT * FROM driver where driver_id = {self.driver_id}")
        data = [element for sub_list in self.cursor.fetchall() for element in sub_list]

        self.cursor.execute(
            f'select ((YEAR(now()) - YEAR(birth_date)) - ((DATE_FORMAT(now(), "00-%m-%d") < DATE_FORMAT(birth_date, "00-%m-%d")))) as age '
            f'from driver where driver_id = {self.driver_id}')
        age = str(self.cursor.fetchall()[0][0])

        intro_layout = QHBoxLayout()
        fio_and_age = QLabel(f'{data[2]} {data[1]}, {age} {"год" if age[-1] == "1" else "года" if age[-1] in "234" else "лет"}')
        intro_layout.addWidget(fio_and_age)

        form_layout = QFormLayout()
        for column in list(zip(columns, data)):
            label = QLabel(column[0])
            data = QLabel(str(column[1]))
            form_layout.addRow(label, data)

        inf_layout = QVBoxLayout()
        inf_layout.addLayout(intro_layout)
        inf_layout.addLayout(form_layout)

        create_dialog = QPushButton('Мои маршруты')
        create_dialog.clicked.connect(self.show_roots)
        inf_layout.addWidget(create_dialog)

        self.setLayout(inf_layout)
        self.center()

    def center(self):
        screen_geometry = QApplication.primaryScreen().geometry()
        window_geometry = self.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.move(window_geometry.topLeft())

    def show_roots(self):
        dialog = QDialog(self)
        dialog_layout = QVBoxLayout()
        dialog.setWindowTitle('Данные о ваших маршрутах')
        close_dialog_button = QPushButton('Закрыть')
        close_dialog_button.clicked.connect(dialog.accept)
        to_excel_button = QPushButton('Вывести в excel')
        to_excel_button.clicked.connect(self.to_excel)
        buttons_layout = QVBoxLayout()

        self.cursor.execute(f"SELECT * FROM routs where driver_id = {self.driver_id}")
        data = self.cursor.fetchall()

        if data:
            dialog.setMinimumSize(600, 250)
            routs_table = QTableWidget()
            routs_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            routs_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            routs_table.verticalHeader().setVisible(False)
            dialog_layout.addWidget(routs_table)
            routs_table.setRowCount(0)
            num_rows = len(data)
            num_cols = len(data[0])

            routs_table.setRowCount(num_rows)
            routs_table.setColumnCount(num_cols)

            column_headers = [description[0] for description in self.cursor.description]
            routs_table.setHorizontalHeaderLabels(column_headers)

            for row in range(num_rows):
                for col in range(num_cols):
                    item = QTableWidgetItem(str(data[row][col]))
                    routs_table.setItem(row, col, item)

            buttons_layout.addWidget(to_excel_button)
        else:
            no_roots = QLabel('Информация по вашим маршрутам отсутствует')
            dialog_layout.addWidget(no_roots)

        buttons_layout.addWidget(close_dialog_button)
        dialog_layout.addLayout(buttons_layout)
        dialog.setLayout(dialog_layout)
        result = dialog.exec()

    def to_excel(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for col_idx, header in enumerate(["id рейса", "Авто", "Водитель", "Дата выезда", "Дата прибытия", "Расстояние", "Пункт назначенния"], start=1):
            column_letter = get_column_letter(col_idx)
            cell = f"{column_letter}1"
            worksheet[cell] = header
            worksheet[cell].alignment = Alignment(horizontal='center')

        self.cursor.execute(
            f'select rout_id, concat(mark, " ", model), concat(female, " ", name), '
            f'departure_date, arrival_date, distance, destination from routs '
            f'inner join auto on auto.auto_id = routs.auto_id '
            f'inner join driver on driver.driver_id = routs.driver_id '
            f'where driver.driver_id = {self.driver_id}')

        data = self.cursor.fetchall()

        for row_idx, row_data in enumerate([list(item) for item in data], start=2):
            for col_idx, value in enumerate(row_data, start=1):
                column_letter = get_column_letter(col_idx)
                cell = f"{column_letter}{row_idx}"
                worksheet[cell] = value
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        workbook.save(f'Мои маршруты.xlsx')
